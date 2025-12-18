#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de génération du fichier Excel Configurateur Mining
Génère un fichier Excel structuré avec 5 onglets à partir du JSON de configuration
"""

import json
import math
import os
from pathlib import Path
from typing import Dict, List, Any, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
        NamedStyle
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERREUR: openpyxl n'est pas installé.")
    print("Installez-le avec: pip install openpyxl")
    exit(1)


# ============================================================================
# CONFIGURATION DES STYLES
# ============================================================================

def get_header_style():
    """Style pour les en-têtes de tableaux"""
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def get_cell_style():
    """Style pour les cellules de données"""
    return {
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'alignment': Alignment(vertical="center")
    }

def get_title_style():
    """Style pour les titres de sections"""
    return {
        'font': Font(bold=True, size=14, color="1F4E78"),
        'fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'alignment': Alignment(horizontal="left", vertical="center")
    }


# ============================================================================
# LECTURE DU JSON
# ============================================================================

def read_json_config(json_path: str) -> Dict[str, Any]:
    """Lit et parse le fichier JSON de configuration"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================================
# ONGLET 1: SETUP_ADMIN
# ============================================================================

def create_setup_admin_sheet(wb: Workbook, config: Dict[str, Any]) -> Dict[str, int]:
    """Crée l'onglet SETUP_ADMIN avec toutes les bibliothèques et paramètres"""
    ws = wb.create_sheet("SETUP_ADMIN")
    
    row = 1
    
    # === BIBLIOTHÈQUE CONTAINERS ===
    ws.merge_cells(f'A{row}:I{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "BIBLIOTHÈQUE CONTAINERS"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["ID", "Nom", "Puissance (MW)", "Courant (A)", "Refroidissement", 
               "CAPEX (USD)", "Longueur (m)", "Largeur (m)", "Fondation"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    containers_start_row = row
    for container in config["equipment_library"]["containers"]:
        cooling = ", ".join(container.get("cooling", []))
        ws.cell(row=row, column=1, value=container["id"])
        ws.cell(row=row, column=2, value=container["name"])
        ws.cell(row=row, column=3, value=container["power_mw"])
        ws.cell(row=row, column=4, value=container.get("current_a", ""))
        ws.cell(row=row, column=5, value=cooling)
        ws.cell(row=row, column=6, value=container["capex_usd"])
        ws.cell(row=row, column=7, value=container["dimensions_m"]["length"])
        ws.cell(row=row, column=8, value=container["dimensions_m"]["width"])
        ws.cell(row=row, column=9, value=container.get("foundation", ""))
        
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    containers_end_row = row - 1
    row += 2
    
    # === BIBLIOTHÈQUE TRANSFORMATEURS ===
    ws.merge_cells(f'A{row}:H{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "BIBLIOTHÈQUE TRANSFORMATEURS"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["ID", "Nom", "Puissance (MW)", "CAPEX (USD)", 
               "Longueur (m)", "Largeur (m)", "Fondation"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    transformers_start_row = row
    for transformer in config["equipment_library"]["transformers"]:
        ws.cell(row=row, column=1, value=transformer["id"])
        ws.cell(row=row, column=2, value=transformer["name"])
        ws.cell(row=row, column=3, value=transformer["power_mw"])
        ws.cell(row=row, column=4, value=transformer["capex_usd"])
        ws.cell(row=row, column=5, value=transformer["dimensions_m"]["length"])
        ws.cell(row=row, column=6, value=transformer["dimensions_m"]["width"])
        ws.cell(row=row, column=7, value=transformer.get("foundation", ""))
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    transformers_end_row = row - 1
    row += 2
    
    # === BIBLIOTHÈQUE POWERBLOCKS ===
    ws.merge_cells(f'A{row}:I{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "BIBLIOTHÈQUE POWERBLOCKS"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["ID", "Nom", "Puissance (MW)", "Nbr Transformateurs", 
               "Redondance", "CAPEX (USD)", "Longueur (m)", "Largeur (m)", "Fondation"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    powerblocks_start_row = row
    for powerblock in config["equipment_library"]["powerblocks"]:
        ws.cell(row=row, column=1, value=powerblock["id"])
        ws.cell(row=row, column=2, value=powerblock["name"])
        ws.cell(row=row, column=3, value=powerblock["power_mw"])
        ws.cell(row=row, column=4, value=powerblock.get("transformers", 0))
        ws.cell(row=row, column=5, value=powerblock.get("redundancy_transformers", 0))
        ws.cell(row=row, column=6, value=powerblock["capex_usd"])
        ws.cell(row=row, column=7, value=powerblock["dimensions_m"]["length"])
        ws.cell(row=row, column=8, value=powerblock["dimensions_m"]["width"])
        ws.cell(row=row, column=9, value=powerblock.get("foundation", ""))
        
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    powerblocks_end_row = row - 1
    row += 2
    
    # === BIBLIOTHÈQUE ÉNERGIE ===
    ws.merge_cells(f'A{row}:D{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "BIBLIOTHÈQUE ÉNERGIE"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Type", "Buffer requis", "OPEX (USD/MWh)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    energy_start_row = row
    for energy_type, energy_data in config["energy_sources"].items():
        ws.cell(row=row, column=1, value=energy_type)
        ws.cell(row=row, column=2, value="Oui" if energy_data["buffer_required"] else "Non")
        ws.cell(row=row, column=3, value=energy_data["opex_usd_per_mwh"])
        
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    energy_end_row = row - 1
    
    # Retourner les informations pour utilisation dans autres onglets
    return {
        "energy_start_row": energy_start_row,
        "energy_end_row": energy_end_row
    }
    row += 2
    
    # === STANDARDS GÉNIE CIVIL ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "STANDARDS GÉNIE CIVIL"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Type", "Nom", "Épaisseur (cm)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    for layer in config["standards"]["civil_work"]["layers"]:
        ws.cell(row=row, column=1, value=layer["type"])
        ws.cell(row=row, column=2, value=layer.get("name", ""))
        ws.cell(row=row, column=3, value=layer["thickness_cm"])
        
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    row += 2
    
    # === RÈGLES ===
    ws.merge_cells(f'A{row}:B{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "RÈGLES"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Règle", "Valeur"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    rules = config["standards"]["rules"]
    rule_mapping = {
        "container_atomic_unit": "Container unité atomique",
        "container_power_mw": "Puissance container (MW)",
        "no_powerblock_below_mw": "Pas de PowerBlock en dessous de (MW)",
        "powerblock_size_mw": "Taille PowerBlock (MW)",
        "transformer_size_mw": "Taille transformateur (MW)",
        "max_containers_per_transformer": "Max containers par transformateur",
        "fuses_location": "Emplacement fusibles",
        "bt_upstream_equipment": "Équipement amont BT"
    }
    
    for rule_key, rule_value in rules.items():
        rule_name = rule_mapping.get(rule_key, rule_key)
        if isinstance(rule_value, bool):
            rule_value = "Oui" if rule_value else "Non"
        ws.cell(row=row, column=1, value=rule_name)
        ws.cell(row=row, column=2, value=rule_value)
        
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)
        row += 1
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12


# ============================================================================
# ONGLET 2: INPUT_PROJECT
# ============================================================================

def create_input_project_sheet(wb: Workbook, config: Dict[str, Any]) -> None:
    """Crée l'onglet INPUT_PROJECT avec les paramètres utilisateur"""
    ws = wb.create_sheet("INPUT_PROJECT")
    
    row = 1
    
    # Titre
    ws.merge_cells(f'A{row}:B{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "PARAMÈTRES DU PROJET"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 2
    
    # Paramètres de base
    labels = [
        ("Nom projet", "C3"),
        ("Pays", "C4"),
        ("Puissance IT cible (MW)", "C5"),
        ("Puissance future (MW)", "C6"),
        ("Type d'énergie", "C7"),
        ("Type mining", "C8"),
        ("Restriction surface", "C9"),
        ("Surface max (m²)", "C10")
    ]
    
    for label, cell_ref in labels:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=":")
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row += 1
    
    # Valeurs par défaut
    project_input = config["project_input"]
    ws["C3"].value = project_input.get("project_name", "")
    ws["C4"].value = project_input.get("country", "")
    ws["C5"].value = project_input.get("power_target_mw", 0)
    ws["C6"].value = project_input.get("future_power_mw", 0)
    ws["C7"].value = project_input.get("energy_type", "grid")
    ws["C8"].value = project_input.get("mining_type", "air")
    ws["C9"].value = "Non" if project_input.get("surface_limit_m2") is None else "Oui"
    ws["C10"].value = project_input.get("surface_limit_m2") or ""
    
    # Validation liste déroulante pour Type d'énergie
    energy_types = list(config["energy_sources"].keys())
    energy_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(energy_types)}"'
    )
    energy_dv.add(ws["C7"])
    ws.add_data_validation(energy_dv)
    
    # Validation liste déroulante pour Type mining
    mining_dv = DataValidation(
        type="list",
        formula1='"air,immersion"'
    )
    mining_dv.add(ws["C8"])
    ws.add_data_validation(mining_dv)
    
    # Validation liste déroulante pour Restriction surface
    restriction_dv = DataValidation(
        type="list",
        formula1='"Oui,Non"'
    )
    restriction_dv.add(ws["C9"])
    ws.add_data_validation(restriction_dv)
    
    row += 2
    
    # Phasage chantier
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "PHASAGE CHANTIER"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Phase", "Puissance (MW)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    phasing = project_input.get("phasing", [])
    for i, phase in enumerate(phasing, start=1):
        ws.cell(row=row, column=1, value=f"Phase {i}")
        ws.cell(row=row, column=2, value=phase.get("power_mw", 0))
        row += 1
    
    # Ligne vide pour ajouter des phases
    ws.cell(row=row, column=1, value="Phase ?").fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 25


# ============================================================================
# ONGLET 3: CALCUL_ENGINE
# ============================================================================

def create_calcul_engine_sheet(wb: Workbook, config: Dict[str, Any], setup_info: Dict[str, int]) -> Dict[str, int]:
    """Crée l'onglet CALCUL_ENGINE avec les formules de calcul"""
    ws = wb.create_sheet("CALCUL_ENGINE")
    
    row = 1
    
    # === CALCULS QUANTITÉS ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "CALCULS QUANTITÉS"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    ws.cell(row=row, column=1, value="Équipement").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Quantité").font = Font(bold=True)
    
    row += 1
    
    rules = config["standards"]["rules"]
    container_power = rules["container_power_mw"]
    transformer_power = rules["transformer_size_mw"]
    powerblock_power = rules["powerblock_size_mw"]
    no_powerblock_below = rules["no_powerblock_below_mw"]
    
    # Containers
    ws.cell(row=row, column=1, value="Containers HD5")
    ws.cell(row=row, column=2, value=f'=ARRONDI.SUP(INPUT_PROJECT!C5/{container_power};0)')
    row += 1
    
    # Transformateurs
    ws.cell(row=row, column=1, value="Transformateurs")
    ws.cell(row=row, column=2, value=f'=ARRONDI.SUP(INPUT_PROJECT!C5/{transformer_power};0)')
    row += 1
    
    # PowerBlocks
    ws.cell(row=row, column=1, value="PowerBlocks")
    ws.cell(row=row, column=2, value=f'=SI(INPUT_PROJECT!C5<{no_powerblock_below};0;ARRONDI.SUP(INPUT_PROJECT!C5/{powerblock_power};0))')
    row += 2
    
    # === CALCULS CAPEX ===
    ws.merge_cells(f'A{row}:E{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "CALCULS CAPEX"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Équipement", "Quantité", "Prix unitaire (USD)", "CAPEX (USD)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    capex_start_row = row
    
    # Stocker les lignes de quantité pour référence
    containers_qty_row = row - 3  # Ligne de quantité containers (B4)
    transformers_qty_row = row - 2  # Ligne de quantité transformateurs (B5)
    powerblocks_qty_row = row - 1  # Ligne de quantité powerblocks (B6)
    
    # Containers CAPEX
    container_id = config["equipment_library"]["containers"][0]["id"]
    container_capex = config["equipment_library"]["containers"][0]["capex_usd"]
    ws.cell(row=row, column=1, value="Containers")
    ws.cell(row=row, column=2, value=f'=B{containers_qty_row}')
    ws.cell(row=row, column=3, value=container_capex)
    ws.cell(row=row, column=4, value=f'=B{row}*C{row}')
    row += 1
    
    # Transformateurs CAPEX
    transformer_id = config["equipment_library"]["transformers"][0]["id"]
    transformer_capex = config["equipment_library"]["transformers"][0]["capex_usd"]
    ws.cell(row=row, column=1, value="Transformateurs")
    ws.cell(row=row, column=2, value=f'=B{transformers_qty_row}')
    ws.cell(row=row, column=3, value=transformer_capex)
    ws.cell(row=row, column=4, value=f'=B{row}*C{row}')
    row += 1
    
    # PowerBlocks CAPEX
    powerblock_id = config["equipment_library"]["powerblocks"][0]["id"]
    powerblock_capex = config["equipment_library"]["powerblocks"][0]["capex_usd"]
    ws.cell(row=row, column=1, value="PowerBlocks")
    ws.cell(row=row, column=2, value=f'=B{powerblocks_qty_row}')
    ws.cell(row=row, column=3, value=powerblock_capex)
    ws.cell(row=row, column=4, value=f'=B{row}*C{row}')
    row += 1
    
    # Génie Civil (à calculer basé sur surface)
    ws.cell(row=row, column=1, value="Génie Civil")
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value='=LAYOUT!B2*100')  # Surface totale * coût/m² (100 USD/m² exemple)
    row += 1
    
    # CAPEX Total
    ws.cell(row=row, column=1, value="CAPEX TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value="")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value=f'=SOMME(D{capex_start_row}:D{row-1})').font = Font(bold=True, size=12)
    capex_total_row = row
    row += 2
    
    # === CALCULS OPEX ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "CALCULS OPEX (ANNUEL)"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    ws.cell(row=row, column=1, value="Élément").font = Font(bold=True)
    ws.cell(row=row, column=2, value="OPEX (USD/an)").font = Font(bold=True)
    
    row += 1
    
    # OPEX Électricité
    # Formule complexe pour récupérer OPEX selon type énergie
    # On utilise INDEX/MATCH pour trouver l'OPEX du type d'énergie sélectionné
    ws.cell(row=row, column=1, value="Électricité")
    # Formule: puissance * 8760 * OPEX selon type énergie
    energy_start = setup_info["energy_start_row"]
    energy_end = setup_info["energy_end_row"]
    energy_col_ref = f'INDEX(SETUP_ADMIN!C{energy_start}:C{energy_end};EQUIV(INPUT_PROJECT!C7;SETUP_ADMIN!A{energy_start}:A{energy_end};0);1)'
    ws.cell(row=row, column=2, value=f'=INPUT_PROJECT!C5*8760*{energy_col_ref}')
    row += 1
    
    # OPEX Maintenance (% CAPEX - exemple 2%)
    opex_maint_row = row
    ws.cell(row=row, column=1, value="Maintenance (2% CAPEX)")
    ws.cell(row=row, column=2, value=f'=D{capex_total_row}*0.02')
    row += 1
    
    # OPEX Total
    opex_total_row = row
    ws.cell(row=row, column=1, value="OPEX TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=f'=SOMME(B{row-2}:B{row-1})').font = Font(bold=True, size=12)
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 5
    
    # Retourner les informations pour utilisation dans autres onglets
    return {
        "capex_containers_row": capex_start_row,
        "capex_transformers_row": capex_start_row + 1,
        "capex_powerblocks_row": capex_start_row + 2,
        "capex_civil_row": capex_start_row + 3,
        "capex_total_row": capex_total_row,
        "opex_electricity_row": row - 2,
        "opex_maintenance_row": row - 1,
        "opex_total_row": opex_total_row
    }


# ============================================================================
# GÉNÉRATION DU LAYOUT AUTOMATIQUE
# ============================================================================

def generate_layout(config: Dict[str, Any], power_target_mw: float) -> List[Dict[str, Any]]:
    """Génère le layout automatique des équipements"""
    layout_data = []
    
    rules = config["standards"]["rules"]
    defaults = config["layout_engine"]["defaults"]
    layout_rules = config["layout_engine"]["rules"]
    
    container_power = rules["container_power_mw"]
    transformer_power = rules["transformer_size_mw"]
    max_containers_per_transformer = rules["max_containers_per_transformer"]
    
    # Dimensions
    container_dim = config["layout_engine"]["elements_dimensions"]["HD5"]
    transformer_dim = config["layout_engine"]["elements_dimensions"]["TR_5MW"]
    
    container_length = container_dim["length"]
    container_width = container_dim["width"]
    transformer_length = transformer_dim["length"]
    transformer_width = transformer_dim["width"]
    
    # Espacements
    container_spacing = defaults["container_spacing_m"]
    row_spacing = defaults["row_spacing_m"]
    road_width = defaults["road_width_m"]
    grass_width = defaults["grass_strip_width_m"]
    
    # Calculs quantités
    num_containers = math.ceil(power_target_mw / container_power)
    num_transformers = math.ceil(power_target_mw / transformer_power)
    
    # Organisation en lignes (3 containers max par transformateur)
    containers_per_row = max_containers_per_transformer
    num_rows = math.ceil(num_containers / containers_per_row)
    
    current_x = road_width  # Commence après la route initiale
    current_y = road_width  # Commence après la route initiale
    
    # Génération des containers et transformateurs
    container_idx = 0
    transformer_idx = 0
    
    for row in range(num_rows):
        row_y = current_y
        
        # Containers de la ligne
        containers_in_row = min(containers_per_row, num_containers - container_idx)
        for col in range(containers_in_row):
            x_pos = current_x + col * (container_length + container_spacing)
            layout_data.append({
                "id": f"HD5-{container_idx + 1}",
                "type": "Container",
                "x": x_pos,
                "y": row_y,
                "rotation": 0,
                "phase": 1,
                "length": container_length,
                "width": container_width
            })
            container_idx += 1
        
        # Transformateur de la ligne (placé après les containers)
        if transformer_idx < num_transformers:
            trans_x = current_x + containers_in_row * (container_length + container_spacing) + container_spacing
            layout_data.append({
                "id": f"TR-{transformer_idx + 1}",
                "type": "Transformateur",
                "x": trans_x,
                "y": row_y,
                "rotation": 0,
                "phase": 1,
                "length": transformer_length,
                "width": transformer_width
            })
            transformer_idx += 1
        
        # Routes autour de chaque container (si règle activée)
        if layout_rules.get("road_around_each_container", False):
            for col in range(containers_in_row):
                cont_x = current_x + col * (container_length + container_spacing)
                # Route nord
                layout_data.append({
                    "id": f"ROAD-N-{row}-{col}",
                    "type": "Route",
                    "x": cont_x - road_width / 2,
                    "y": row_y - road_width / 2,
                    "rotation": 0,
                    "phase": 1,
                    "length": container_length + road_width,
                    "width": road_width
                })
                # Route sud
                layout_data.append({
                    "id": f"ROAD-S-{row}-{col}",
                    "type": "Route",
                    "x": cont_x - road_width / 2,
                    "y": row_y + container_width + road_width / 2,
                    "rotation": 0,
                    "phase": 1,
                    "length": container_length + road_width,
                    "width": road_width
                })
                # Route est
                layout_data.append({
                    "id": f"ROAD-E-{row}-{col}",
                    "type": "Route",
                    "x": cont_x + container_length + road_width / 2,
                    "y": row_y - road_width / 2,
                    "rotation": 0,
                    "phase": 1,
                    "length": road_width,
                    "width": container_width + road_width
                })
                # Route ouest
                layout_data.append({
                    "id": f"ROAD-W-{row}-{col}",
                    "type": "Route",
                    "x": cont_x - road_width / 2,
                    "y": row_y - road_width / 2,
                    "rotation": 0,
                    "phase": 1,
                    "length": road_width,
                    "width": container_width + road_width
                })
        
        # Route autour de la ligne (si règle activée)
        if layout_rules.get("road_around_each_row", False) and row == 0:
            # Route ouest de la ligne
            row_width = containers_in_row * (container_length + container_spacing) - container_spacing
            layout_data.append({
                "id": f"ROAD-ROW-W-{row}",
                "type": "Route",
                "x": current_x - road_width,
                "y": row_y - road_width / 2,
                "rotation": 0,
                "phase": 1,
                "length": road_width,
                "width": container_width + road_width + (transformer_width if transformer_idx <= num_transformers else 0)
            })
        
        # Mise à jour Y pour la ligne suivante
        max_width_in_row = max(container_width, transformer_width if transformer_idx <= num_transformers else 0)
        current_y += max_width_in_row + row_spacing
        
        # Gazon entre les lignes (si règle activée et pas dernière ligne)
        if layout_rules.get("grass_between_rows", False) and row < num_rows - 1:
            grass_y = row_y + max_width_in_row + road_width
            layout_data.append({
                "id": f"GRASS-{row}",
                "type": "Gazon",
                "x": current_x - road_width,
                "y": grass_y,
                "rotation": 0,
                "phase": 1,
                "length": row_width + road_width * 2,
                "width": grass_width
            })
            current_y += grass_width
    
    return layout_data


# ============================================================================
# ONGLET 4: LAYOUT
# ============================================================================

def create_layout_sheet(wb: Workbook, config: Dict[str, Any]) -> None:
    """Crée l'onglet LAYOUT avec le tableau des positions"""
    ws = wb.create_sheet("LAYOUT")
    
    row = 1
    
    # Surface totale (sera calculée)
    ws.cell(row=row, column=1, value="Surface totale utilisée (m²):").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=SOMMEPROD(G6:G1000;H6:H1000)')
    
    # Vérification limite de surface (si définie)
    row += 1
    ws.cell(row=row, column=1, value="Surface limite (m²):").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=SI(INPUT_PROJECT!C9="Oui";INPUT_PROJECT!C10;"Non limité")')
    
    row += 1
    ws.cell(row=row, column=1, value="Alerte dépassement:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=SI(ET(INPUT_PROJECT!C9="Oui";B2>B3);"⚠ DÉPASSEMENT";"OK")')
    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row += 2
    
    # Tableau Layout
    headers = ["ID Élément", "Type", "X (m)", "Y (m)", "Rotation (°)", "Phase", "Longueur (m)", "Largeur (m)", "Alerte"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    
    # Les données seront générées dynamiquement via une formule ou seront pré-remplies
    # Pour l'instant, on crée les en-têtes et on laisse les cellules vides/modifiables
    # Note: Les données du layout seront générées et remplies après création de INPUT_PROJECT
    # car elles dépendent de la puissance cible
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20


# ============================================================================
# ONGLET 5: GRAPHIQUES
# ============================================================================

def create_graphiques_sheet(wb: Workbook, config: Dict[str, Any], calcul_info: Dict[str, int]) -> None:
    """Crée l'onglet GRAPHIQUES avec les données et graphiques"""
    ws = wb.create_sheet("GRAPHIQUES")
    
    row = 1
    
    # === RÉPARTITION CAPEX ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "RÉPARTITION CAPEX"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Élément", "CAPEX (USD)", "Part (%)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    capex_pie_start_row = row
    
    # Références aux valeurs CAPEX de CALCUL_ENGINE
    containers_capex_row = row
    ws.cell(row=row, column=1, value="Containers")
    ws.cell(row=row, column=2, value=f'=CALCUL_ENGINE!D{calcul_info["capex_containers_row"]}')
    row += 1
    
    ws.cell(row=row, column=1, value="Transformateurs")
    ws.cell(row=row, column=2, value=f'=CALCUL_ENGINE!D{calcul_info["capex_transformers_row"]}')
    row += 1
    
    ws.cell(row=row, column=1, value="PowerBlocks")
    ws.cell(row=row, column=2, value=f'=CALCUL_ENGINE!D{calcul_info["capex_powerblocks_row"]}')
    row += 1
    
    ws.cell(row=row, column=1, value="Génie Civil")
    ws.cell(row=row, column=2, value=f'=CALCUL_ENGINE!D{calcul_info["capex_civil_row"]}')
    row += 1
    
    total_row = row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value='=SOMME(B6:B9)').font = Font(bold=True)
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
    # Ajouter formule pour les pourcentages
    ws.cell(row=containers_capex_row, column=3, value='=SI($B$10<>0;B6/$B$10;0)')
    ws.cell(row=containers_capex_row+1, column=3, value='=SI($B$10<>0;B7/$B$10;0)')
    ws.cell(row=containers_capex_row+2, column=3, value='=SI($B$10<>0;B8/$B$10;0)')
    ws.cell(row=containers_capex_row+3, column=3, value='=SI($B$10<>0;B9/$B$10;0)')
    capex_total_row_graph = row
    row += 3
    
    # Graphique camembert CAPEX
    pie = PieChart()
    pie.title = "Répartition CAPEX"
    labels = Reference(ws, min_col=1, min_row=capex_pie_start_row, max_row=capex_pie_start_row+3)
    data = Reference(ws, min_col=2, min_row=capex_pie_start_row-1, max_row=capex_pie_start_row+3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 10
    pie.width = 15
    ws.add_chart(pie, "E2")
    
    # === CAPEX vs OPEX ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "CAPEX vs OPEX ANNUEL"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Type", "CAPEX (USD)", "OPEX (USD/an)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    row += 1
    comparison_start_row = row
    
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=2, value=f'=CALCUL_ENGINE!D{calcul_info["capex_total_row"]}')
    ws.cell(row=row, column=3, value=f'=CALCUL_ENGINE!B{calcul_info["opex_total_row"]}')
    row += 1
    
    # Graphique barres CAPEX vs OPEX
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "CAPEX vs OPEX Annuel"
    bar.y_axis.title = "Montant (USD)"
    bar.x_axis.title = "Type"
    data = Reference(ws, min_col=2, min_row=comparison_start_row-1, max_col=3, max_row=comparison_start_row)
    cats = Reference(ws, min_col=1, min_row=comparison_start_row, max_row=comparison_start_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 10
    bar.width = 15
    ws.add_chart(bar, f"E{comparison_start_row + 5}")
    
    row += 3
    
    # === ÉVOLUTION PAR PHASE ===
    ws.merge_cells(f'A{row}:C{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "ÉVOLUTION PAR PHASE"
    for key, value in get_title_style().items():
        setattr(title_cell, key, value)
    
    row += 1
    headers = ["Phase", "Puissance (MW)", "CAPEX (USD)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        for key, value in get_header_style().items():
            setattr(cell, key, value)
    
    # Les données par phase seront dynamiques basées sur INPUT_PROJECT
    # Pour l'instant, on crée la structure
    
    # Ajuster largeurs colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18


# ============================================================================
# FONCTION PRINCIPALE
# ============================================================================

def populate_layout_sheet(wb: Workbook, config: Dict[str, Any]) -> None:
    """Remplit l'onglet LAYOUT avec les données générées"""
    ws = wb["LAYOUT"]
    
    # Utiliser une puissance par défaut pour générer le layout initial
    # L'utilisateur pourra modifier les positions ensuite
    power_target = config["project_input"].get("power_target_mw", 50)  # Exemple: 50 MW
    
    layout_data = generate_layout(config, power_target)
    
    # Remplir les données
    start_row = 6  # Après les en-têtes (ligne 5) et infos surface
    for idx, item in enumerate(layout_data, start=start_row):
        ws.cell(row=idx, column=1, value=item["id"])
        ws.cell(row=idx, column=2, value=item["type"])
        ws.cell(row=idx, column=3, value=item["x"])  # Modifiable
        ws.cell(row=idx, column=4, value=item["y"])  # Modifiable
        ws.cell(row=idx, column=5, value=item["rotation"])  # Modifiable
        ws.cell(row=idx, column=6, value=item["phase"])
        ws.cell(row=idx, column=7, value=item["length"])
        ws.cell(row=idx, column=8, value=item["width"])
        
        # Alerte chevauchement (formule simplifiée - à améliorer selon besoin)
        # Cette formule vérifie si un élément chevauche avec les autres
        # Note: C'est une version simplifiée, pour un vrai système il faudrait une VBA ou une logique plus complexe
        ws.cell(row=idx, column=9, value='')  # Alerte vide pour l'instant
        
        # Mettre en évidence les cellules modifiables (X, Y, Rotation)
        for col in [3, 4, 5]:
            cell = ws.cell(row=idx, column=col)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Appliquer le style aux autres cellules
        for col in [1, 2, 6, 7, 8, 9]:
            cell = ws.cell(row=idx, column=col)
            for key, value in get_cell_style().items():
                setattr(cell, key, value)


def main():
    """Fonction principale"""
    # Chemins
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    json_path = project_root.parent / "mining_configurator_full_v2.json"
    output_path = project_root / "mining_configurator.xlsx"
    
    print(f"Lecture du JSON: {json_path}")
    if not json_path.exists():
        print(f"ERREUR: Fichier JSON introuvable: {json_path}")
        return
    
    # Lire configuration
    config = read_json_config(str(json_path))
    
    # Créer workbook
    print("Création du fichier Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer feuille par défaut
    
    # Créer les onglets
    print("Création de l'onglet SETUP_ADMIN...")
    setup_info = create_setup_admin_sheet(wb, config)
    
    print("Création de l'onglet INPUT_PROJECT...")
    create_input_project_sheet(wb, config)
    
    print("Création de l'onglet CALCUL_ENGINE...")
    calcul_info = create_calcul_engine_sheet(wb, config, setup_info)
    
    print("Création de l'onglet LAYOUT...")
    create_layout_sheet(wb, config)
    
    print("Création de l'onglet GRAPHIQUES...")
    create_graphiques_sheet(wb, config, calcul_info)
    
    print("Remplissage du layout initial...")
    populate_layout_sheet(wb, config)
    
    # Sauvegarder
    print(f"Sauvegarde du fichier: {output_path}")
    wb.save(str(output_path))
    print("✓ Fichier Excel généré avec succès!")
    print(f"\nEmplacement: {output_path}")


if __name__ == "__main__":
    main()

